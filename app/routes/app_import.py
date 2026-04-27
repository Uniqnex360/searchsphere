import tempfile
import csv
import os
from datetime import date, datetime
from io import StringIO
from typing import Optional
from openpyxl import load_workbook
from sqlalchemy import desc, func
from sqlmodel import select
from sqlmodel.ext.asyncio.session import AsyncSession
from fastapi import APIRouter, File, UploadFile, HTTPException, Depends, Query

from app.models import APPImport
from app.database import get_session
from app.services import ImportType, CeleryTaskStatus


router = APIRouter()

UPLOAD_DIR = "/app/uploads"


COLUMNS = {
    "sku",
    "gtin",
    "ean",
    "upc",
    "taxonomy",
    "country_of_origin",
    "warranty",
    "weight_unit",
    "dimension_unit",
    "currency",
    "stock_status",
    "vendor_name",
    "vendor_sku",
    "short_description",
    "long_description",
    "meta_title",
    "meta_description",
    "search_keywords",
    "certification",
    "safety_standard",
    "hazardous_material",
    "prop65_warning",
    "Product Type",
    "industry_name",
    "brand",
    "industry_name",
    "review",
}


@router.post("/import/product/")
async def upload_products_file(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session),
):

    filename = file.filename.lower()

    # =====================================================
    # ✅ FILE TYPE VALIDATION
    # =====================================================
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        raise HTTPException(400, "Only CSV or XLSX files allowed")

    # =====================================================
    # ✅ HANDLE CSV (YOUR ORIGINAL LOGIC)
    # =====================================================
    if filename.endswith(".csv"):

        # 1. Read header safely
        header_bytes = b""

        while True:
            chunk = await file.read(1024)
            if not chunk:
                break

            header_bytes += chunk

            if b"\n" in chunk:
                break

        if not header_bytes:
            raise HTTPException(400, "Empty file")

        # 2. UTF-8 validation
        try:
            header_text = header_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            raise HTTPException(
                400,
                "File must be UTF-8 encoded CSV. Please save as 'CSV UTF-8'",
            )

        # 3. Extract header
        header_line = header_text.splitlines()[0]

        # 4. Parse header
        try:
            reader = csv.reader(StringIO(header_line))
            headers = next(reader, None)

            if not headers:
                raise HTTPException(400, "Invalid CSV header")

            headers = [h.strip() for h in headers]

        except Exception:
            raise HTTPException(400, "Invalid CSV format")

        # 5. Validate columns
        missing = COLUMNS - set(headers)
        if missing:
            raise HTTPException(400, f"Missing columns: {missing}")

        # 6. Reset pointer
        file.file.seek(0)

        # 7. Save CSV (streaming)
        with tempfile.NamedTemporaryFile(
            delete=False, suffix=".csv", dir=UPLOAD_DIR
        ) as tmp:
            while True:
                chunk = file.file.read(1024 * 1024)
                if not chunk:
                    break
                tmp.write(chunk)

            temp_path = tmp.name

    # =====================================================
    # ✅ HANDLE XLSX (STREAMING CONVERSION)
    # =====================================================
    else:
        # 1. Save XLSX first (streaming)
        with tempfile.NamedTemporaryFile(
            delete=False, suffix=".xlsx", dir=UPLOAD_DIR
        ) as tmp_xlsx:
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                tmp_xlsx.write(chunk)

            xlsx_path = tmp_xlsx.name

        try:
            # 2. Open in read-only mode (VERY IMPORTANT)
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active

            # 3. Read header row
            try:
                header_row = next(ws.iter_rows(max_row=1, values_only=True))
            except StopIteration:
                raise HTTPException(400, "Empty Excel file")

            headers = [str(h).strip() if h is not None else "" for h in header_row]

            if not headers:
                raise HTTPException(400, "Invalid Excel header")

            # 4. Validate columns
            missing = COLUMNS - set(headers)
            if missing:
                raise HTTPException(400, f"Missing columns: {missing}")

            # 5. Convert XLSX → CSV (stream row-by-row)
            with tempfile.NamedTemporaryFile(
                delete=False,
                suffix=".csv",
                dir=UPLOAD_DIR,
                mode="w",
                newline="",
                encoding="utf-8",
            ) as tmp_csv:

                writer = csv.writer(tmp_csv)

                for row in ws.iter_rows(values_only=True):
                    writer.writerow(["" if v is None else v for v in row])

                temp_path = tmp_csv.name

        finally:
            # 6. Cleanup XLSX temp file
            if os.path.exists(xlsx_path):
                os.remove(xlsx_path)

    # =====================================================
    # ✅ COMMON FLOW (UNCHANGED)
    # =====================================================
    obj = APPImport(module_type=ImportType.PRODUCT)

    session.add(obj)
    await session.commit()
    await session.refresh(obj)

    from app.tasks import import_products_task

    task = import_products_task.delay(temp_path, obj.id)

    obj.task_id = task.id
    await session.commit()

    return {
        "message": "Import started",
        "task_id": task.id,
        "import_id": obj.id,
    }


@router.get("/import/list/")
async def get_import_list(
    session: AsyncSession = Depends(get_session),
    status: Optional[CeleryTaskStatus] = Query(None, description="Filter by status"),
    module_type: Optional[ImportType] = Query(
        None, description="Filter by module type"
    ),
    start_date: Optional[datetime] = Query(
        None, description="Filter from this date (YYYY-MM-DD)"
    ),
    end_date: Optional[datetime] = Query(
        None, description="Filter to this date (YYYY-MM-DD)"
    ),
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
):
    """Returns paginated list of APPImport entries with date filters."""

    # 1. Build Base Filter Logic
    filters = []
    if status:
        filters.append(APPImport.status == status)
    if module_type:
        filters.append(APPImport.module_type == module_type)

    # Date Filtering Logic
    if start_date:
        filters.append(APPImport.created_at >= start_date)
    if end_date:
        # Note: If created_at is a DateTime, you might want to ensure end_date
        # includes the full day (e.g., < end_date + 1 day)
        filters.append(APPImport.created_at <= end_date)

    # 2. Optimized Count Query (Executes on DB, not in Python memory)
    count_query = select(func.count()).select_from(APPImport)
    if filters:
        count_query = count_query.where(*filters)

    total_count_res = await session.execute(count_query)
    total_count = total_count_res.scalar() or 0

    # 3. Fetch Paginated Results
    query = select(APPImport).order_by(desc(APPImport.created_at))
    if filters:
        query = query.where(*filters)

    result = await session.execute(query.offset(offset).limit(limit))
    imports = result.scalars().all()

    data = [
        {
            "task_id": imp.task_id,
            "module_type": imp.module_type,
            "status": imp.status,
            "rows": imp.rows,
            "meta_data": imp.meta_data,
            "result": imp.result,
            "error": imp.error,
            "created_at": imp.created_at,
            "completed_at": imp.completed_at,
        }
        for imp in imports
    ]

    return {"total": total_count, "limit": limit, "offset": offset, "data": data}
